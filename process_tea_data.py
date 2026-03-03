"""
Tea Gardens Rainfall Data Processor v2
----------------------------------------
Processes all daily Excel files and generates:
  1. Summary overview
  2. Daily Averages (avg rainfall across all gardens per day - NOT a sum)
  3. Garden Statistics
  4. Monthly Summary (per year)
  5. Yearly Reports - one sheet per garden, month-by-month across years
  6. Top Rain Events
  7. Harvest Calendar (seasonal rainfall guide)

Usage:
    python process_tea_data_v2.py <input_folder> <output_file.xlsx>

Example:
    python process_tea_data_v2.py ~/Downloads/obubu-data/ tea_report_v2.xlsx
"""

import sys
import os
import glob
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER    = "1F4E79"
C_SUBHEADER = "2E75B6"
C_ALT       = "D6E4F0"
C_GREEN     = "375623"
C_GREEN_LT  = "E2EFDA"
C_AMBER     = "7F6000"
C_AMBER_LT  = "FFEB9C"
C_RED       = "9C0006"
C_RED_LT    = "FFC7CE"
C_WHITE     = "FFFFFF"
C_GREY      = "F2F2F2"

MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec"]

# Tea harvest seasons (Uji region, Japan)
HARVEST_NOTES = {
    1:  "Winter dormancy – minimal growth",
    2:  "Late winter – bud development begins",
    3:  "Pre-season – first flush approaching",
    4:  "🌱 First Flush (Ichibancha) – most prized harvest",
    5:  "🌱 First Flush continues – critical rainfall period",
    6:  "🌿 Second Flush (Nibancha) – summer harvest",
    7:  "☀️ Summer – heat stress risk if low rainfall",
    8:  "☀️ Late summer – plant recovery period",
    9:  "🍂 Autumn harvest (Sanbancha) – third flush",
    10: "🍂 Autumn – plant hardening before winter",
    11: "Late autumn – growth slows",
    12: "Winter dormancy – rest period",
}


def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(cell, bg=C_HEADER, size=11):
    cell.font = Font(bold=True, color=C_WHITE, size=size, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def data_style(cell, bold=False, bg=None, align="right", color="000000", size=10):
    cell.font = Font(bold=bold, color=color, size=size, name="Arial")
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()

def rain_colour(mm):
    """Return (text_colour, bg_colour) based on rainfall intensity."""
    if mm >= 100:  return C_RED,   C_RED_LT
    if mm >= 50:   return C_AMBER, C_AMBER_LT
    if mm >= 10:   return C_GREEN, C_GREEN_LT
    return "000000", None


# ── Data Loading ──────────────────────────────────────────────────────────────

def parse_date(d):
    return datetime.strptime(str(int(d)), "%Y%m%d%H%M")


def load_all_files(folder):
    pattern = os.path.join(folder, "**", "*TeaGardensData.xlsx")
    files = sorted(glob.glob(pattern, recursive=True))
    if not files:
        files = sorted(glob.glob(os.path.join(folder, "*TeaGardensData.xlsx")))
    print(f"Found {len(files)} files…")

    records = []
    garden_meta = {}

    for i, fpath in enumerate(files):
        if i % 50 == 0:
            print(f"  Processing file {i+1}/{len(files)}…")
        try:
            xl = pd.read_excel(fpath, sheet_name=None, header=0)
        except Exception as e:
            print(f"  Warning: could not read {fpath}: {e}")
            continue

        if "Tea Gardens" in xl:
            for _, row in xl["Tea Gardens"].iterrows():
                name = row.get("Tea Garden Name")
                if pd.notna(name) and name not in garden_meta:
                    garden_meta[name] = {
                        "japanese": row.get("茶畑名前", ""),
                        "lat": row.get("Latitude"),
                        "lon": row.get("Longitude"),
                    }

        for sheet, df in xl.items():
            if sheet == "Tea Gardens" or df is None or df.empty:
                continue
            try:
                df.columns = ["Type", "Date", "Rainfall"]
                df = df.dropna(subset=["Date", "Rainfall"])
                df["datetime"] = df["Date"].apply(parse_date)
                df["date"]     = df["datetime"].dt.date
                df["garden"]   = sheet
                records.append(df[["garden", "datetime", "date", "Rainfall"]])
            except Exception:
                continue

    if not records:
        raise ValueError("No data found. Check the folder path.")

    raw = pd.concat(records, ignore_index=True)
    raw["Rainfall"] = pd.to_numeric(raw["Rainfall"], errors="coerce").fillna(0)
    return raw, garden_meta


# ── Aggregations ──────────────────────────────────────────────────────────────

def build_daily(raw):
    """Daily total per garden, plus average across all gardens for that day."""
    per_garden = (
        raw.groupby(["garden", "date"])["Rainfall"]
        .sum()
        .reset_index()
        .rename(columns={"Rainfall": "daily_mm"})
    )
    dt = pd.to_datetime(per_garden["date"])
    per_garden["year"]  = dt.dt.year
    per_garden["month"] = dt.dt.month

    # Average across gardens per day (representative rainfall for that day)
    daily_avg = (
        per_garden.groupby("date")["daily_mm"]
        .mean()
        .reset_index()
        .rename(columns={"daily_mm": "avg_mm"})
    )
    return per_garden, daily_avg


def build_monthly_by_year(per_garden):
    """For each garden: total rainfall per (year, month)."""
    return (
        per_garden.groupby(["garden", "year", "month"])["daily_mm"]
        .sum()
        .reset_index()
        .rename(columns={"daily_mm": "monthly_mm"})
    )


def build_garden_stats(per_garden):
    g = per_garden.groupby("garden")["daily_mm"]
    stats = pd.DataFrame({
        "Total Rainfall (mm)":      g.sum(),
        "Avg Daily Rainfall (mm)":  g.mean(),
        "Max Single Day (mm)":      g.max(),
        "Days with Rain":           (per_garden["daily_mm"] > 0).groupby(per_garden["garden"]).sum(),
        "Days Recorded":            g.count(),
        "Heavy Rain Days (>50mm)":  (per_garden["daily_mm"] > 50).groupby(per_garden["garden"]).sum(),
        "Extreme Days (>100mm)":    (per_garden["daily_mm"] > 100).groupby(per_garden["garden"]).sum(),
    }).reset_index().rename(columns={"garden": "Tea Garden"})
    stats["Rain Day %"] = (stats["Days with Rain"] / stats["Days Recorded"] * 100).round(1)
    return stats.sort_values("Total Rainfall (mm)", ascending=False)


def build_top_events(daily_avg, n=30):
    return daily_avg.sort_values("avg_mm", ascending=False).head(n).copy()


# ── Sheet Writers ──────────────────────────────────────────────────────────────

def write_cover(wb, per_garden, garden_meta):
    ws = wb.create_sheet("📋 Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32

    ws.merge_cells("A1:B1")
    ws["A1"] = "🍵  Obubu Tea Gardens — Rainfall Report"
    ws["A1"].font = Font(bold=True, size=20, color=C_HEADER, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y')}"
    ws["A2"].font = Font(size=11, color="666666", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")

    years = sorted(per_garden["year"].unique())
    info = [
        ("Days Recorded",    per_garden["date"].nunique()),
        ("Date Range",       f"{per_garden['date'].min()} → {per_garden['date'].max()}"),
        ("Years Covered",    ", ".join(str(y) for y in years)),
        ("Tea Gardens",      per_garden["garden"].nunique()),
    ]
    for r, (k, v) in enumerate(info, start=4):
        c = ws.cell(r, 1, k)
        c.font = Font(bold=True, size=11, name="Arial")
        v_cell = ws.cell(r, 2, str(v))
        v_cell.font = Font(size=11, name="Arial")
        v_cell.alignment = Alignment(horizontal="right")

    ws.cell(10, 1, "Rainfall Intensity Guide").font = Font(bold=True, size=11, color=C_HEADER, name="Arial")
    guide = [
        ("≥ 100mm / day", "Extreme", C_RED_LT, C_RED),
        ("≥ 50mm / day",  "Heavy",   C_AMBER_LT, C_AMBER),
        ("≥ 10mm / day",  "Moderate",C_GREEN_LT, C_GREEN),
        ("< 10mm / day",  "Light",   None, "000000"),
    ]
    for i, (threshold, label, bg, fg) in enumerate(guide, start=11):
        ws.cell(i, 1, threshold).font = Font(size=10, name="Arial")
        c = ws.cell(i, 2, label)
        c.font = Font(bold=True, color=fg, size=10, name="Arial")
        if bg:
            c.fill = PatternFill("solid", start_color=bg)

    ws.cell(17, 1, "Tea Gardens Included").font = Font(bold=True, size=11, color=C_HEADER, name="Arial")
    for i, g in enumerate(sorted(per_garden["garden"].unique()), start=18):
        meta = garden_meta.get(g, {})
        ws.cell(i, 1, g).font = Font(size=10, name="Arial")
        ws.cell(i, 2, meta.get("japanese", "")).font = Font(size=10, name="Arial")

    ws.cell(18 + per_garden["garden"].nunique() + 1, 1,
            "NOTE: 'Daily Rainfall' figures show the AVERAGE across all gardens, not a sum.").font = \
        Font(italic=True, size=9, color="666666", name="Arial")


def write_daily_avg(wb, daily_avg, per_garden):
    """
    Daily sheet: average rainfall across all gardens (representative day reading)
    plus each garden's individual total for reference.
    """
    ws = wb.create_sheet("📅 Daily Averages")
    ws.sheet_view.showGridLines = False

    gardens = sorted(per_garden["garden"].unique())
    pivot = (
        per_garden.pivot_table(index="date", columns="garden", values="daily_mm", aggfunc="sum")
        .reindex(columns=gardens)
        .fillna(0)
    )
    pivot = pivot.join(daily_avg.set_index("date")["avg_mm"])
    pivot = pivot.sort_index()

    # Headers
    ws.cell(1, 1, "Date")
    header_style(ws.cell(1, 1))
    ws.column_dimensions["A"].width = 14

    ws.cell(1, 2, "Avg Across All Gardens (mm)")
    header_style(ws.cell(1, 2), bg=C_GREEN)
    ws.column_dimensions["B"].width = 28

    for ci, g in enumerate(gardens, start=3):
        header_style(ws.cell(1, ci, g), bg=C_SUBHEADER, size=9)
        ws.column_dimensions[get_column_letter(ci)].width = 13

    for ri, (date, row) in enumerate(pivot.iterrows(), start=2):
        bg_row = C_ALT if ri % 2 == 0 else None
        ws.cell(ri, 1, str(date))
        data_style(ws.cell(ri, 1), bg=bg_row, align="center")

        avg = row["avg_mm"]
        tc, bg = rain_colour(avg)
        c = ws.cell(ri, 2, round(avg, 2))
        c.number_format = '#,##0.00'
        c.font = Font(bold=True, color=tc, size=10, name="Arial")
        c.fill = PatternFill("solid", start_color=bg) if bg else (PatternFill("solid", start_color=bg_row) if bg_row else PatternFill())
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = thin_border()

        for ci, g in enumerate(gardens, start=3):
            val = row.get(g, 0)
            cell = ws.cell(ri, ci, round(val, 2))
            cell.number_format = '#,##0.00'
            data_style(cell, bg=bg_row)

    ws.freeze_panes = "A2"
    ws.cell(1, 1).value = "Date"


def write_garden_stats(wb, stats):
    ws = wb.create_sheet("🌿 Garden Statistics")
    ws.sheet_view.showGridLines = False
    cols = list(stats.columns)
    widths = [25, 20, 22, 20, 14, 14, 22, 20, 12]
    for ci, (col, w) in enumerate(zip(cols, widths), start=1):
        header_style(ws.cell(1, ci, col))
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(stats.itertuples(index=False), start=2):
        bg = C_ALT if ri % 2 == 0 else None
        for ci, val in enumerate(row, start=1):
            c = ws.cell(ri, ci)
            if isinstance(val, float):
                c.value = round(val, 2)
                c.number_format = '#,##0.00'
            else:
                c.value = val
            data_style(c, bg=bg, align="left" if ci == 1 else "right")
    ws.freeze_panes = "A2"


def write_yearly_garden_sheets(wb, monthly_by_year):
    """
    One sheet per garden. Rows = months, Columns = years.
    Shows total rainfall per month per year.
    Colour coded by intensity.
    """
    gardens = sorted(monthly_by_year["garden"].unique())
    years   = sorted(monthly_by_year["year"].unique())

    for garden in gardens:
        safe_name = garden[:28]  # Excel sheet name limit
        ws = wb.create_sheet(f"🍃 {safe_name}")
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells(f"A1:{get_column_letter(len(years)+3)}1")
        ws["A1"] = f"{garden}  —  Monthly Rainfall by Year (mm)"
        ws["A1"].font = Font(bold=True, size=13, color=C_HEADER, name="Arial")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Column headers: Month | Jan..Dec | Year Total
        ws.cell(2, 1, "Month")
        header_style(ws.cell(2, 1))
        ws.column_dimensions["A"].width = 12

        for ci, yr in enumerate(years, start=2):
            header_style(ws.cell(2, ci, str(yr)), bg=C_SUBHEADER)
            ws.column_dimensions[get_column_letter(ci)].width = 12

        avg_col = len(years) + 2
        header_style(ws.cell(2, avg_col, "Month Avg"), bg=C_GREEN)
        ws.column_dimensions[get_column_letter(avg_col)].width = 13

        harvest_col = avg_col + 1
        header_style(ws.cell(2, harvest_col, "🌱 Harvest Season Notes"), bg="375623")
        ws.column_dimensions[get_column_letter(harvest_col)].width = 42

        gdf = monthly_by_year[monthly_by_year["garden"] == garden]

        for month in range(1, 13):
            row = month + 2  # row 3 = Jan, row 14 = Dec
            bg = C_ALT if month % 2 == 0 else None

            ws.cell(row, 1, MONTH_NAMES[month-1])
            data_style(ws.cell(row, 1), bold=True, bg=bg, align="center")

            year_vals = []
            for ci, yr in enumerate(years, start=2):
                match = gdf[(gdf["year"] == yr) & (gdf["month"] == month)]
                val = match["monthly_mm"].values[0] if not match.empty else 0.0
                year_vals.append(val)
                c = ws.cell(row, ci, round(val, 1))
                c.number_format = '#,##0.0'
                tc, cell_bg = rain_colour(val / 30)  # monthly ÷ ~30 days for daily equiv
                if cell_bg:
                    c.fill = PatternFill("solid", start_color=cell_bg)
                    c.font = Font(bold=(val > 200), color=tc, size=10, name="Arial")
                else:
                    data_style(c, bg=bg)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="right", vertical="center")

            # Month average across years
            avg_val = sum(year_vals) / len(year_vals) if year_vals else 0
            c_avg = ws.cell(row, avg_col, round(avg_val, 1))
            c_avg.number_format = '#,##0.0'
            c_avg.font = Font(bold=True, size=10, name="Arial")
            data_style(c_avg, bold=True, bg=bg)
            c_avg.border = thin_border()

            # Harvest note
            note_cell = ws.cell(row, harvest_col, HARVEST_NOTES.get(month, ""))
            note_cell.font = Font(italic=True, size=9, name="Arial")
            note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if bg:
                note_cell.fill = PatternFill("solid", start_color=bg)
            note_cell.border = thin_border()

        # Year totals footer
        footer = 15
        ws.cell(footer, 1, "YEAR TOTAL")
        data_style(ws.cell(footer, 1), bold=True, bg=C_HEADER)
        ws.cell(footer, 1).font = Font(bold=True, color=C_WHITE, size=10, name="Arial")
        ws.cell(footer, 1).fill = PatternFill("solid", start_color=C_HEADER)

        for ci, yr in enumerate(years, start=2):
            cl = get_column_letter(ci)
            ws.cell(footer, ci, f"=SUM({cl}3:{cl}14)")
            ws.cell(footer, ci).number_format = '#,##0.0'
            ws.cell(footer, ci).font = Font(bold=True, color=C_WHITE, size=10, name="Arial")
            ws.cell(footer, ci).fill = PatternFill("solid", start_color=C_HEADER)
            ws.cell(footer, ci).border = thin_border()

        avg_cl = get_column_letter(avg_col)
        ws.cell(footer, avg_col, f"=SUM({avg_cl}3:{avg_cl}14)")
        ws.cell(footer, avg_col).number_format = '#,##0.0'
        ws.cell(footer, avg_col).font = Font(bold=True, color=C_WHITE, size=10, name="Arial")
        ws.cell(footer, avg_col).fill = PatternFill("solid", start_color=C_HEADER)
        ws.cell(footer, avg_col).border = thin_border()

        ws.freeze_panes = "B3"
        ws.row_dimensions[2].height = 28


def write_all_gardens_monthly(wb, monthly_by_year):
    """
    Cross-garden view: for each year+month, show all gardens side by side.
    Useful for spotting which gardens got more/less rain in key months.
    """
    years   = sorted(monthly_by_year["year"].unique())
    gardens = sorted(monthly_by_year["garden"].unique())

    ws = wb.create_sheet("📊 All Gardens × Month")
    ws.sheet_view.showGridLines = False

    # Headers
    ws.cell(1, 1, "Year")
    header_style(ws.cell(1, 1))
    ws.column_dimensions["A"].width = 8

    ws.cell(1, 2, "Month")
    header_style(ws.cell(1, 2))
    ws.column_dimensions["B"].width = 10

    for ci, g in enumerate(gardens, start=3):
        header_style(ws.cell(1, ci, g), bg=C_SUBHEADER, size=9)
        ws.column_dimensions[get_column_letter(ci)].width = 13

    avg_col = len(gardens) + 3
    header_style(ws.cell(1, avg_col, "Garden Avg (mm)"), bg=C_GREEN)
    ws.column_dimensions[get_column_letter(avg_col)].width = 17

    ri = 2
    for yr in years:
        for month in range(1, 13):
            bg = C_ALT if ri % 2 == 0 else None

            ws.cell(ri, 1, str(yr))
            data_style(ws.cell(ri, 1), bg=bg, align="center")

            ws.cell(ri, 2, MONTH_NAMES[month-1])
            data_style(ws.cell(ri, 2), bold=True, bg=bg, align="center")

            vals = []
            for ci, g in enumerate(gardens, start=3):
                match = monthly_by_year[
                    (monthly_by_year["garden"] == g) &
                    (monthly_by_year["year"] == yr) &
                    (monthly_by_year["month"] == month)
                ]
                val = match["monthly_mm"].values[0] if not match.empty else 0.0
                vals.append(val)
                c = ws.cell(ri, ci, round(val, 1))
                c.number_format = '#,##0.0'
                tc, cell_bg = rain_colour(val / 30)
                if cell_bg and val > 0:
                    c.fill = PatternFill("solid", start_color=cell_bg)
                    c.font = Font(color=tc, size=9, name="Arial")
                else:
                    data_style(c, bg=bg, size=9)
                c.border = thin_border()
                c.alignment = Alignment(horizontal="right", vertical="center")

            avg_val = sum(vals) / len(vals) if vals else 0
            c_avg = ws.cell(ri, avg_col, round(avg_val, 1))
            c_avg.number_format = '#,##0.0'
            data_style(c_avg, bold=True, bg=bg)

            ri += 1

    ws.freeze_panes = "A2"


def write_top_events(wb, top_events):
    ws = wb.create_sheet("⛈️ Top Rain Events")
    ws.sheet_view.showGridLines = False

    headers = ["Rank", "Date", "Avg Rainfall (mm)", "Intensity", "Harvest Season Context"]
    widths  = [8, 14, 20, 24, 42]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        header_style(ws.cell(1, ci, h))
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(top_events.itertuples(index=False), start=2):
        bg = C_ALT if ri % 2 == 0 else None
        mm = row.avg_mm
        month = pd.to_datetime(row.date).month

        if mm >= 100:   intensity = "🌊 Extreme (≥100mm avg)"
        elif mm >= 50:  intensity = "🌧️ Very Heavy (≥50mm avg)"
        elif mm >= 20:  intensity = "🌦️ Heavy (≥20mm avg)"
        else:           intensity = "🌂 Moderate"

        vals = [ri - 1, str(row.date), round(mm, 2), intensity, HARVEST_NOTES.get(month, "")]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(ri, ci, val)
            data_style(c, bg=bg, align="center" if ci in (1,2) else "left")
            if ci == 3:
                c.number_format = '#,##0.00'
                tc, cell_bg = rain_colour(mm)
                if cell_bg:
                    c.fill = PatternFill("solid", start_color=cell_bg)
                    c.font = Font(bold=True, color=tc, size=10, name="Arial")


def write_harvest_calendar(wb, monthly_by_year):
    """Summary view: avg rainfall per month across all gardens, all years — harvest planning view."""
    ws = wb.create_sheet("🌱 Harvest Calendar")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "Seasonal Rainfall Guide — All Gardens, All Years"
    ws["A1"].font = Font(bold=True, size=14, color=C_HEADER, name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["Month", "Avg Monthly Rainfall (mm)", "Avg Daily Equiv (mm/day)", "Harvest Season Notes", "Rainfall Assessment"]
    widths  = [12, 26, 26, 44, 22]
    for ci, (h, w) in enumerate(zip(headers, widths), start=1):
        header_style(ws.cell(2, ci, h))
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Aggregate: mean monthly rainfall across all gardens and years
    monthly_mean = (
        monthly_by_year.groupby("month")["monthly_mm"]
        .mean()
        .reset_index()
    )

    for ri, row in enumerate(monthly_mean.itertuples(index=False), start=3):
        m = row.month
        avg_monthly = row.monthly_mm
        avg_daily   = avg_monthly / 30

        if avg_daily >= 10:   assessment = "🟢 Well watered"
        elif avg_daily >= 5:  assessment = "🟡 Adequate"
        elif avg_daily >= 2:  assessment = "🟠 Moderate – monitor"
        else:                 assessment = "🔴 Low – potential stress"

        bg = C_ALT if ri % 2 == 0 else None
        vals = [MONTH_NAMES[m-1], round(avg_monthly, 1), round(avg_daily, 2),
                HARVEST_NOTES.get(m, ""), assessment]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(ri, ci, val)
            data_style(c, bg=bg, align="left" if ci > 2 else "center")
            if ci == 2:
                c.number_format = '#,##0.0'
            if ci == 3:
                c.number_format = '#,##0.00'


# ── Main ──────────────────────────────────────────────────────────────────────

def main(input_folder, output_path):
    print("Loading files…")
    raw, garden_meta = load_all_files(input_folder)
    print(f"Loaded {len(raw):,} readings | {raw['garden'].nunique()} gardens | {raw['date'].nunique()} days")

    print("Aggregating…")
    per_garden, daily_avg = build_daily(raw)
    monthly_by_year       = build_monthly_by_year(per_garden)
    stats                 = build_garden_stats(per_garden)
    top_events            = build_top_events(daily_avg, n=30)

    print("Writing Excel report…")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_cover(wb, per_garden, garden_meta)
    write_daily_avg(wb, daily_avg, per_garden)
    write_garden_stats(wb, stats)
    write_all_gardens_monthly(wb, monthly_by_year)
    write_yearly_garden_sheets(wb, monthly_by_year)
    write_top_events(wb, top_events)
    write_harvest_calendar(wb, monthly_by_year)

    wb.save(output_path)
    print(f"\n✅ Done! Report saved to: {output_path}")
    print(f"   Sheets created: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"     • {s}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])