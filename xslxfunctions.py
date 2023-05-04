import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from os.path import join, isfile
import os
from datetime import date

date_string = str(date.today())


def read_xlsx_file(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Create a list to hold all the dictionaries
    data = []
    
    # Get the header row and use it as keys for each row
    keys = [cell.value for cell in ws[1]]
    
    # Loop over the rows starting from the second row
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Create a dictionary using the keys and row values
        row_dict = dict(zip(keys, row))
        # Add the dictionary to the list
        data.append(row_dict)
    
    return data

def createxlsxworkbook(date_string='', filename_prefix='', sheet_names=[], xlsx_folder_path=''):
    # Create file path
    xlsx_wb_name = f"{date_string}{filename_prefix}.xlsx"
    xlsx_file_path_full = os.path.join(xlsx_folder_path, xlsx_wb_name)

    # Check if workbook exists
    if os.path.isfile(xlsx_file_path_full):
        print(f"Workbook '{xlsx_wb_name}' already exists at '{xlsx_file_path_full}'.")
    else:
        # Create workbook and sheets
        wb = Workbook()
        for sheet_name in sheet_names:
            wb.create_sheet(sheet_name)

        # Save workbook to file path
        wb.save(xlsx_file_path_full)

        print(f"Workbook '{xlsx_wb_name}' created at '{xlsx_file_path_full}'.")

    return [xlsx_file_path_full, xlsx_wb_name]


def write_xlsx_worksheet(list_of_dict, sheet_name='', xlsx_file_loc=''):
    # Open workbook and worksheet
    wb = load_workbook(xlsx_file_loc)

    # Check if sheet exists
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Sheet '{sheet_name}' already exists in workbook '{xlsx_file_loc}'.")
    else:
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        print(f"Sheet '{sheet_name}' created in workbook '{xlsx_file_loc}'.")

    # Get existing column headings
    existing_headings = []
    if ws[1][0].value is not None:
        existing_headings = [cell.value for cell in ws[1]]

    # Write headings if there are none in the first row
    if not existing_headings:
        existing_headings = list(list_of_dict[0].keys())
        for col_num, heading in enumerate(existing_headings, start=1):
            ws.cell(row=1, column=col_num).value = heading
            ws.cell(row=1, column=col_num).font = Font(bold=True)

    # Write data to matching columns or new columns
    for cus_dict in list_of_dict:
        row_num = ws.max_row + 1
        for key, value in cus_dict.items():
            if key not in existing_headings:
                # Add new column heading
                new_col = get_column_letter(ws.max_column + 1)
                ws[f"{new_col}1"].value = key
                ws[f"{new_col}1"].font = Font(bold=True)
                existing_headings.append(key)

            # Write data to correct column
            col_num = existing_headings.index(key) + 1
            ws.cell(row=row_num, column=col_num, value=value)


    # Save workbook
    wb.save(xlsx_file_loc)

def copy_sheet_contents(src_file_path, src_sheet_name, dest_file_path, dest_sheet_name):
    # Load source workbook and sheet
    src_wb = load_workbook(filename=src_file_path)
    src_ws = src_wb[src_sheet_name]

    # Load destination workbook and sheet
    dest_wb = load_workbook(filename=dest_file_path)
    dest_ws = dest_wb[dest_sheet_name]

    # Copy source sheet contents to destination sheet
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(row)

    existing_headings = []
    if dest_ws[1][0].value is not None:
        existing_headings = [cell.value for cell in dest_ws[1]]
    for col_num, heading in enumerate(existing_headings, start=1):
        dest_ws.cell(row=1, column=col_num).value = heading
        dest_ws.cell(row=1, column=col_num).font = Font(bold=True)

    # Save destination workbook
    dest_wb.save(dest_file_path)

def rename_sheet(workbook_path, old_sheet_name, new_sheet_name):
    # Load workbook
    wb = load_workbook(filename=workbook_path)

    # Get sheet by old name
    old_sheet = wb[old_sheet_name]

    # Rename sheet
    old_sheet.title = new_sheet_name

    # Save workbook
    wb.save(workbook_path)

def remove_empty_sheets(workbook_path):
    # Load workbook
    wb = load_workbook(filename=workbook_path)

    # Remove empty sheets
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        if sheet.max_row == 0 and sheet.max_column == 0:
            wb.remove(sheet)

    # Save workbook
    wb.save(workbook_path)

def rename_active_sheet(workbook_path, new_sheet_name):
    # Load workbook and get the active sheet
    wb = load_workbook(filename=workbook_path)
    ws = wb.active

    # Rename the active sheet
    ws.title = new_sheet_name

    # Save the workbook
    wb.save(workbook_path)

def delete_xlsx_rows(sheet_names=[], rows=[], xlsx_file_loc=''):
    # Open workbook
    wb = load_workbook(xlsx_file_loc)

    # Loop over sheet names
    for sheet_name in sheet_names:
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in workbook '{xlsx_file_loc}'.")
            continue

        # Open worksheet
        ws = wb[sheet_name]

        # Loop over rows to delete
        for row_num in rows:
            # Check if row exists
            if row_num < 1 or row_num > ws.max_row:
                print(f"Row '{row_num}' not found in sheet '{sheet_name}' of workbook '{xlsx_file_loc}'.")
                continue

            # Delete row
            ws.delete_rows(row_num)

    # Save workbook
    wb.save(xlsx_file_loc)

def delete_rows_with_duplicates(sheet_name='', col_name='', xlsx_file_loc=''):
    # Open workbook and worksheet
    wb = load_workbook(xlsx_file_loc)

    # Check if sheet exists
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in workbook '{xlsx_file_loc}'.")
        return

    ws = wb[sheet_name]

    # Get column index from column name
    col_index = None
    for cell in ws[1]:
        if cell.value == col_name:
            col_index = cell.column
            break

    if col_index is None:
        print(f"Column '{col_name}' not found in sheet '{sheet_name}' of workbook '{xlsx_file_loc}'.")
        return

    # Create set to store unique values
    unique_values = set()

    # Loop over rows and check for duplicates
    for row in reversed(range(1, ws.max_row + 1)):
        cell_value = ws.cell(row=row, column=col_index).value
        if cell_value in unique_values:
            # Delete entire row if duplicate found
            ws.delete_rows(row)
            print(f"Row {row} in sheet '{sheet_name}' of workbook '{xlsx_file_loc}' deleted due to duplicate value '{cell_value}' in column '{col_name}'.")
        else:
            unique_values.add(cell_value)

    # Save workbook
    wb.save(xlsx_file_loc)

def get_sheet_names(xlsx_file_loc=''):
    wb = load_workbook(xlsx_file_loc)
    sheetnames = wb.sheetnames
    return sheetnames