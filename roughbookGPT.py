import requests
from dotenv import load_dotenv
from os import path
import os

root = path.dirname(path.abspath(__file__))
dotenvpath = str(root) + '/weather-checker.env'
load_dotenv(dotenv_path=dotenvpath)

yahooapiclientid = os.getenv('yahooapiclientid')

def getweather(lat, lon, appid, outputformat):
    url = f"https://map.yahooapis.jp/weather/V1/place?coordinates={lon},{lat}&appid={appid}&output={outputformat}"
    response = requests.get(url)
    responsejson = response.json()
    print(responsejson)  # Print the entire JSON response
    temp_list = responsejson["Feature"][0]["Property"]["WeatherList"]["Weather"][0]["Temperature"]
    return temp_list


temperature = getweather(34.784835, 135.874422, yahooapiclientid, 'json')
print(f"The current temperature is {temperature}°C.")








# GPT refactored writetoxlsx code

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

def write_xlsx_worksheet(list_of_dict, sheet_name='', xlsx_file_loc=''):
    wb = load_workbook(xlsx_file_loc)
    ws = get_or_create_worksheet(wb, sheet_name)

    existing_headings = set(cell.value for cell in ws[1])
    write_headings(ws, existing_headings, list_of_dict)

    if isinstance(list_of_dict, list):
        first_row_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        write_data_to_columns(ws, list_of_dict, existing_headings, first_row_cells)

    elif isinstance(list_of_dict, dict):
        first_row_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        write_data_to_columns(ws, [list_of_dict], existing_headings, first_row_cells)

    wb.save(xlsx_file_loc)

def get_or_create_worksheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        return wb.create_sheet(sheet_name)

def write_headings(ws, existing_headings, list_of_dict):
    if not existing_headings:
        existing_headings = list(list_of_dict[0].keys())
        for col_num, heading in enumerate(existing_headings, start=1):
            ws.cell(row=1, column=col_num).value = heading
            ws.cell(row=1, column=col_num).font = Font(bold=True)

def write_data_to_columns(ws, list_of_dict, existing_headings, first_row_cells):
    for cus_dict in list_of_dict:
        row_num = ws.max_row + 1
        for key, value in cus_dict.items():
            if key not in existing_headings:
                # Add new column heading
                col_num = ws.max_column + 1
                ws.cell(row=1, column=col_num).value = key
                ws.cell(row=1, column=col_num).font = Font(bold=True)
                existing_headings.add(key)

            # Write data to correct column
            col_num = column_index_from_string(first_row_cells.__next__().column) + list(existing_headings).index(key)
            ws.cell(row=row_num, column=col_num, value=value)

            # Write data in batches of 100 rows
            if row_num % 100 == 0:
                ws.parent.flush()

# optimised delete rows function

from openpyxl import load_workbook

def delete_rows_with_duplicates(sheet_name='', col_name='', xlsx_file_loc=''):
    # Open workbook and worksheet
    wb = load_workbook(xlsx_file_loc)

    # Check if sheet exists
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in workbook '{xlsx_file_loc}'.")
        return

    ws = wb[sheet_name]

    # Get column index from column name
    col_index = ws.cell(row=1, column=1, value=col_name).column

    if col_index is None:
        print(f"Column '{col_name}' not found in sheet '{sheet_name}' of workbook '{xlsx_file_loc}'.")
        return

    # Create dictionary to store unique values
    unique_values = {}

    # Loop over rows in reverse order and check for duplicates
    for row in reversed(range(2, ws.max_row + 1)):
        cell_value = ws.cell(row=row, column=col_index).value
        if cell_value in unique_values:
            # Delete entire row if duplicate found
            ws.delete_rows(row)
            print(f"Row {row} in sheet '{sheet_name}' of workbook '{xlsx_file_loc}' deleted due to duplicate value '{cell_value}' in column '{col_name}'.")
        else:
            unique_values[cell_value] = True

    # Save workbook
    wb.save(xlsx_file_loc)
